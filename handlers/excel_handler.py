import openpyxl
from os.path import abspath


def get_tasks_4_problems_from_excel(path='handlers/materials_for_studying/task_4/task_4.xlsx'):
    users_accents_word = []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        users_accents_word.append({'correct_word': str(row[0]), 'incorrect_word': str(row[1])})
    return users_accents_word


def get_tasks_problems_from_excel(number_of_task: int):
    try:
        if number_of_task == 4:
            return get_tasks_4_problems_from_excel()
        path = rf'handlers/materials_for_studying/task_{number_of_task}/task_{number_of_task}.xlsx'
        users_problems = []
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            users_problems.append({'num_in_column': str(row[0]), 'description': str(row[1]), 'text_of_task': str(row[2]),
                                   'correct_answers': str(row[3]), 'explanation': str(row[4]),
                                   'video_with_explanation': str(row[5]), 'time_code_of_video': str(row[6]),
                                   'number_of_theory': str(row[7])})
        return users_problems
    except Exception as ex:
        print(f'Ошибка при получении задания {number_of_task}','//->', ex)
